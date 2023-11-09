import sys
import os

from Ui_file_select_widget import Ui_Dialog
from PySide6.QtWidgets import (
    QApplication, 
    QWidget, 
    QFileDialog, 
    QGridLayout,
    QPushButton, 
    QDialog,
)

import openpyxl
import datetime
# from time import strftime

class CustomDialog(QDialog):
    run_path = ""
    return_state = 0
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

        self.setWindowTitle("选择以什么方式保存")
        self.ui.dialog_open_button.clicked.connect(self.open_handler)
        self.ui.dialog_new_button.clicked.connect(self.new_handler)
        self.ui.dialog_cancel_button.clicked.connect(self.cancel_handler)

        self.run_path = os.getcwd()

    def open_handler(self):
        print("用户点击了'打开'按钮")
        self.return_state = 1
        self.accept()  # 关闭对话框

    def cancel_handler(self):
        self.return_state = 0
        print("用户点击了'取消'按钮")
        self.reject()  # 关闭对话框

    def new_handler(self):
        self.return_state = 2
        print("用户点击了'新建'按钮")
        self.accept()  # 关闭对话框

    def create_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly  # 可以更改为QFileDialog.WriteOnly以仅允许写入
        file_dialog = QFileDialog(self)
        file_path, _ = file_dialog.getSaveFileName(self, '新建文件', '', 'Text Files (*.xlsx);;All Files (*)', options=options)

        if file_path:
            try:
                with open(file_path, 'w') as file:
                    # 文件新建,应该初始化文件
                    pass
                # 由于excel文件有特定格式,直接创建,会导致格式有问题
                # 删掉再创建
                os.remove(file_path) 
                # 创建一个新的工作簿对象
                workbook = openpyxl.Workbook()
                # 选择要写入数据的工作表
                sheet = workbook.active
                # 写入初始化数据,记录锚点的数据为B1,首次数据存储地点为A3开始
                data_to_write = [
                    ["当前写入数据的起始行号:", 3, "备注：手动更改起始行号，可以指定软件从哪行开始写入"],
                    ["写入时间", "序列号", "链接", "原始数据"],
                ]
                for row_data in data_to_write:
                    sheet.append(row_data)
                workbook.save(file_path)
                    
                print(f'文件已成功创建：{file_path}')
                return file_path
            except Exception as e:
                print(f'创建文件时发生错误：{e}')
                return ""
            
    def Open_file(self):
        dialog = QFileDialog(self)
        dialog.setDirectory(self.run_path)
        dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        dialog.setNameFilter("表格 (*.xlsx)")
        dialog.setViewMode(QFileDialog.ViewMode.List)
        if dialog.exec():
            filenames = dialog.selectedFiles()
            if filenames:
                return filenames[0]
            else:
                return ""
    
def Open_file(widegt):
    dialog_main = CustomDialog(widegt)
    if dialog_main.exec():
        match dialog_main.return_state:
            case 1:
                return dialog_main.Open_file()
            case 2:
                return dialog_main.create_file()

def Save_in_file(file_addr, txt = "错误", uri = "错误", byte = "错误"):
    
    try:
        # 打开.xlsx文件
        workbook = openpyxl.load_workbook(file_addr)
        # 选择要读取数据的工作表
        sheet = workbook['Sheet']  # 这里使用工作表的名称，你可以根据实际情况修改
        # 读取到第一行数据,B1记录了最后写入的表格行号
        history_data = sheet["B1"].value
        
        # # 从行对象中读取单元格的值
        # history_row = row[1]
        # 选择要写入数据的工作表
        sheet = workbook.active
        #向excel中写入对应的value
        # 获取当前时间
        now=datetime.datetime.now()
        sheet.cell(row=history_data, column=1).value = now.strftime("%Y-%m-%d %H:%M:%S")
        sheet.cell(row=history_data, column=2).value = txt
        sheet.cell(row=history_data, column=3).value = uri
        sheet.cell(row=history_data, column=4).value = byte

        # 用来记录历史的锚点数据更新
        history_data += 1
        sheet["B1"] = history_data

        # 保存文件
        workbook.save(file_addr)
        return  True
        print(f'写入文件成功')
    except Exception as e:
        print(f'操作文件时发生错误：{e}')
        return  False

# 测试
class MainWindow(QWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.setWindowTitle('PyQt File Dialog')
        self.setGeometry(100, 100, 300, 150)

        layout = QGridLayout()
        self.setLayout(layout)

        # file selection
        file_browser_btn = QPushButton('Browse')
        file_browser_btn.clicked.connect(self.open_dialog)
        layout.addWidget(file_browser_btn, 2 ,0)

        self.show()
        

    def open_dialog(self):
        temp_file_addr = Open_file(self)
        print(temp_file_addr)
        Save_in_file(temp_file_addr, "小伙子", "真", "帅")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec())