#-*- coding:utf8 -*-
import os
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl import utils

#将指定的xls文件转换为xlsx格式
def xls_to_xlsx(src_file_path,dst_file_path):
    """
    将指定的xls文件转换为xlsx格式
    :param src_file_path: xls文件路径
    :param dst_file_path: xlsx文件路径
    """
    #打开xls文件
    wb_xls = xlrd.open_workbook(f'{src_file_path}')
    #获取所有sheet名称
    names = wb_xls.sheet_names()
    #我们以第一个sheet内容复制到xlsx文件为例，如有多个sheet要复制加循环即可
    #获取xls文件第一个sheet对象
    ws_xls = wb_xls.sheet_by_name(names[0])
    #查看第一个sheet有多少行和列数据
    nrows = ws_xls.nrows
    ncols = ws_xls.ncols
    print(f"xls转xlsx中，列数：{nrows}, 行数{ncols}")
    
    #读取xls第一个sheet内容存到写入列表中
    sheet_msg_xls = []
    for i in range(nrows):
        sheet_msg_xls.append(ws_xls.row_values(i))
    print(f"转换后的列表长度：{len(sheet_msg_xls)}")
    #创建xlsx文件并写入到对应文件中，一般文件同名,sheet名称一致
    wb_xlsx = Workbook()
    #在xlsx创建同名sheet
    ws_xlxs = wb_xlsx.create_sheet(f'{names[0]}')
    #把xls sheet内容写入到 xlsx同名sheet中
    for row in sheet_msg_xls:
        #一行一行的写入
        ws_xlxs.append(row)
    new_sheet_names = wb_xlsx.sheetnames
    del wb_xlsx[new_sheet_names[0]]
    #保存xlsx格式文件
    wb_xlsx.save(f'{dst_file_path}')

def ColAligement(ws,col):
    stringList =[]
    maxLength =0
    
    for cell in ws[utils.get_column_letter(col)]:
        length =0
        try:
            value = float(cell.value)
        except Exception as e:
            value = cell.value
        else:
            # 浮点数有时候会很长，精确到小数点后几位
            value = round(value, 3)
        # print(value)
        for char in str(value):
            if ord(char)> 256:
                length += 1.7
            else:length += 1
        if length > maxLength:
            maxLength = length
    # print(maxLength)
    ws.column_dimensions[utils.get_column_letter(col)].width = maxLength+4

def auto_adjust_column_width(sheet):
    # print(sheet)
    for col in range(1, sheet.max_column + 1):  # 修正：包括最后一列
        ColAligement(sheet, col)

def auto_adjust_row_height(sheet):
    for row in range(1, sheet.max_row + 1):  # 修正：包括最后一行
        sheet.row_dimensions[row].height = 18

def add_grid_lines(sheet):
    # 创建一个边框样式
    thin_border = Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    
    # 为所有单元格设置边框
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

def run(in_file_path = "", out_file_path = "", title = ""):
    # print(f"{in_file_path}, {out_file_path}, {title}")
    stash_in_file = "temp.xlsx"
    if os.path.splitext(in_file_path)[1] == ".xls":
        xls_to_xlsx(in_file_path, stash_in_file)
    else:
        stash_in_file = in_file_path
    in_sheet_name = os.path.splitext(os.path.basename(in_file_path))[0]
    #载入工作表
    in_wb = load_workbook(filename = stash_in_file, data_only=True)
    try:
        in_sheet = in_wb[in_sheet_name]
    except Exception as e:
        print(f"未找到与文件名同名工作薄，改为直接载入第一个工作薄，错误信息：{e}")
        in_sheet = in_wb.worksheets[0]
    else:
        pass        
    # 旧表筛选重复供应商并保存到数据中
    row_index = 3
    column_index = 1
    data = {}  # 初始化data为一个空字典
    while True:
        cell_value = in_sheet.cell(row=row_index, column=column_index).value
        if cell_value is None:
            break  # 如果当前单元格为空，结束循环
        if cell_value in data:  # 如果data中已有此值
            data[cell_value].append(row_index)  # 将当前行号加入对应的列表
        else:
            data[cell_value] = [row_index]  # 如果是第一次遇到该值，创建新的列表
        row_index += 1

    # 创建新表
    new_wb = Workbook()
    # 新表填入各种工作薄
    new_sheets = []
    for sheet_name in data:
        new_sheets.append(new_wb.create_sheet(sheet_name))

    max_column = in_sheet.max_column +1
    #处理各个工作薄
    for new_sheet in new_sheets:
        # 表格内容里的分类，从旧表复制到新表
        for col in range(1, max_column):
            source_cell = in_sheet.cell(row=2, column=col)
            new_sheet.cell(row=2, column=col).value = source_cell.value
        # 从数据结构中获取到同一名称的多个行数
        in_sheet_row = 3
        for write_row in data[new_sheet.title]:
            # 复制整行数据
            for write_col in range(1, max_column):
                # print(f"{write_row}, {write_col}")
                # 获取源单元格的值
                source_cell = in_sheet.cell(row=write_row, column=write_col)
                # 将值复制到目标单元格
                new_sheet.cell(row=in_sheet_row, column=write_col).value = source_cell.value
            in_sheet_row += 1

        max_row = new_sheet.max_row +1

        # 找到列数
        col = 0
        for value in new_sheet[2]:
            col += 1
            if value.value == "数量":
                #在 数量 前两个单元格插入合计
                new_sheet.cell(row = max_row, column = col-2).value = "合计"
                break
        # 计算本列和
        sum = 0.0
        for row in range(3, max_row):
            value = new_sheet.cell(row = row, column = col).value
            sum += float(value)
        new_sheet.cell(row = max_row, column = col).value = sum
        
        # 计算价税合计
        col = 0
        for value in new_sheet[2]:
            col += 1
            if value.value == "价税合计":
                break
        # 计算本列和
        sum = 0.0
        for row in range(3, max_row):
            value = new_sheet.cell(row = row, column = col).value
            sum += float(value)
        new_sheet.cell(row = max_row, column = col).value = sum

        max_row = new_sheet.max_row +1
        max_column = new_sheet.max_column +1

        #整个表格左对齐
        for row in range(2, max_row):
            for col in range(1, max_column):
                center_alignment = Alignment(horizontal='left', vertical='center')
                new_sheet.cell(row=row, column=col).alignment  = center_alignment
        # 添加网格线
        add_grid_lines(new_sheet)

        auto_adjust_column_width(new_sheet)
        auto_adjust_row_height(new_sheet)

        # 标题
        new_sheet.cell(row=1, column=1).value = new_sheet.title + title
        # 合并单元格
        new_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column-1)
        # 居中显示
        center_alignment = Alignment(horizontal='center', vertical='center')
        new_sheet.cell(row=1, column=1).alignment  = center_alignment

    #重命名各个工作薄
    for new_sheet in new_sheets:
        new_sheet.title = new_sheet.title + title
        
    del_sheet_name = new_wb.sheetnames
    # 删除第一个表，因为第一个时默认生成的空表
    del new_wb[del_sheet_name[0]]

    # 设置工作表的打印缩放比例，使其内容适合一页。
    for sheet in new_wb.worksheets:
        sheet.page_setup.fitToWidth = 1
        # sheet.page_setup.fitToHeight = 1

    new_wb.save(out_file_path)


if __name__ == "__main__":
    run("./2025-01月份采购入库明细(1).xls", "2025-01月份采购入库明细处理完成.xlsx", "sdf")
