#-*- coding:utf8 -*-
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

#将指定的xls文件转换为xlsx格式
def xls_to_xlsx(src_file_path,dst_file_path):
    """
    将指定的xls文件转换为xlsx格式
    :param src_file_path: xls文件路径
    :param dst_file_path: xlsx文件路径
    """
    #打开xls文件
    wb_xls = xlrd.open_workbook(f'{src_file_path}')
    #获取所有sheet名称
    names = wb_xls.sheet_names()
    #我们以第一个sheet内容复制到xlsx文件为例，如有多个sheet要复制加循环即可
    #获取xls文件第一个sheet对象
    ws_xls = wb_xls.sheet_by_name(names[0])
    #查看第一个sheet有多少行和列数据
    nrows = ws_xls.nrows
    ncols = ws_xls.ncols
    print(nrows,ncols)
    
    #读取xls第一个sheet内容存到写入列表中
    sheet_msg_xls = []
    for i in range(nrows):
        sheet_msg_xls.append(ws_xls.row_values(i))
    print(len(sheet_msg_xls))
    #创建xlsx文件并写入到对应文件中，一般文件同名,sheet名称一致
    wb_xlsx = Workbook()
    #在xlsx创建同名sheet
    ws_xlxs = wb_xlsx.create_sheet(f'{names[0]}')
    #把xls sheet内容写入到 xlsx同名sheet中
    for row in sheet_msg_xls:
        #一行一行的写入
        ws_xlxs.append(row)
    #保存xlsx格式文件
    wb_xlsx.save(f'{dst_file_path}')    


src_file_path = "2025-01月份采购入库明细.xls"
new_file_name = "2025-01月份采购入库明细.xlsx"

xls_to_xlsx(src_file_path, new_file_name)

wb = load_workbook(filename = new_file_name)
sheet_ranges = wb["2025-01月份采购入库明细"]
data = {}
row_index = 3
column_index = 1

data = {}  # 初始化data为一个空字典
while True:
    cell_value = sheet_ranges.cell(row=row_index, column=column_index).value
    if cell_value is None:
        break  # 如果当前单元格为空，结束循环
    if cell_value in data:  # 如果data中已有此值
        data[cell_value].append(row_index)  # 将当前行号加入对应的列表
    else:
        data[cell_value] = [row_index]  # 如果是第一次遇到该值，创建新的列表
    row_index += 1

new_file = Workbook()

for sheet_name in data:
    new_sheet = new_file.create_sheet(sheet_name)

    max_column = sheet_ranges.max_column

    new_sheet.cell(row=1, column=1).value = new_file_name
    new_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    # 居中显示
    center_alignment = Alignment(horizontal='center', vertical='center')
    new_sheet.cell(row=1, column=1).alignment  = center_alignment
    # 表头
    for col in range(1, max_column + 1):
        source_cell = sheet_ranges.cell(row=2, column=col)
        new_sheet.cell(row=2, column=col).value = source_cell.value

    target_row = 3
    for row in data[sheet_name]:
        # print(f"{sheet_name}, {row}")
        # 复制整行数据
        for col in range(1, max_column + 1):
            # print(f"{row}, {col}")
            # 获取源单元格的值
            source_cell = sheet_ranges.cell(row=row, column=col)
            # print(source_cell.value)
            # 将值复制到目标单元格
            new_sheet.cell(row=target_row, column=col).value = source_cell.value
        target_row += 1
        # print(new_sheet)
if "Sheet" in new_file.sheetnames:
    del new_file["Sheet"]


new_file.save('balances.xlsx')
